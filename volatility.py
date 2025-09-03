import sys
import argparse
from pathlib import Path
from typing import Optional, Dict, Tuple, List
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from scipy import stats
sys.path.append('/home/eric/Documents/growi')
from orderbook_utils import parse_orderbook_blob, validate_required_columns, REQUIRED_CSV_COLUMNS, _normalize_number
from API import get_leverage_info_from_file, LEVERAGE_DATA_FILE
DEFAULT_PERCENTILES = [1, 10, 25, 50]
DEFAULT_MAX_LEVELS = 8
EXCEL_FILENAME = 'volatility_analysis.xlsx'
PLOTS_DIR = Path("plots")
PLOTS_DIR.mkdir(exist_ok=True)

def calculate_log_returns(prices: pd.Series) -> pd.Series:
    prices = prices[prices > 0]
    log_returns = np.log(prices / prices.shift(1))
    return log_returns.dropna()

def calculate_volatility_metrics(log_returns: pd.Series) -> Dict[str, float]:
    if len(log_returns) < 2:
        return {'3min_vol': 0.0, '5min_vol': 0.0, '10min_vol': 0.0, '30min_vol': 0.0, '60min_vol': 0.0, '90min_vol': 0.0, '1day_vol': 0.0, '1year_vol': 0.0, 'sample_size': len(log_returns), 'mean_return': 0.0}
    vol_5min = log_returns.std()
    return {'3min_vol': vol_5min * np.sqrt(3 / 5), '5min_vol': vol_5min, '10min_vol': vol_5min * np.sqrt(10 / 5), '30min_vol': vol_5min * np.sqrt(30 / 5), '60min_vol': vol_5min * np.sqrt(60 / 5), '90min_vol': vol_5min * np.sqrt(90 / 5), '1day_vol': vol_5min * np.sqrt(1440 / 5), '1year_vol': vol_5min * np.sqrt(525600 / 5), 'sample_size': len(log_returns), 'mean_return': log_returns.mean()}

def fetch_leverage_info(symbol: str) -> Dict:
    return get_leverage_info_from_file(symbol, LEVERAGE_DATA_FILE)

def get_effective_leverage_for_usd_amount(usd_amount: float, leverage_tiers: List[Tuple[float, float, int]]) -> int:
    if not leverage_tiers:
        return 0
    for lower_bound, upper_bound, leverage in leverage_tiers:
        if lower_bound <= usd_amount < upper_bound:
            return leverage
    return leverage_tiers[-1][2] if leverage_tiers else 0

def analyze_csv(csv_path: str, symbol_filters: Optional[list]=None, sep: str=',') -> Dict:
    try:
        df = pd.read_csv(csv_path, sep=sep, engine='python')
    except Exception as e:
        raise SystemExit(f'Error reading CSV: {e}')
    validate_required_columns(df.columns.tolist(), set(REQUIRED_CSV_COLUMNS))
    if symbol_filters:
        df = df[df['symbol'].isin(symbol_filters)]
        if len(df) == 0:
            raise SystemExit(f'No data found for symbols: {symbol_filters}')
        found_symbols = df['symbol'].unique()
        missing_symbols = set(symbol_filters) - set(found_symbols)
    if 'ts' in df.columns:
        df['ts'] = pd.to_datetime(df['ts'], format='mixed')
        df = df.sort_values('ts')
    results = {}
    for symbol in df['symbol'].unique():
        sdf = df[df['symbol'] == symbol].copy()
        prices = sdf['base_price'].apply(_normalize_number)
        prices = pd.Series(prices, dtype=float)
        if 'ts' in sdf.columns:
            prices.index = sdf['ts']
        log_returns = calculate_log_returns(prices)
        vol_metrics = calculate_volatility_metrics(log_returns)
        leverage_info = fetch_leverage_info(symbol)
        if leverage_info['status'] == 'error':
            print("error fetching leverage info for", symbol, leverage_info['error'])
        results[symbol] = {'df': sdf, 'volatility_metrics': vol_metrics, 'leverage_info': leverage_info, 'price_stats': {'min_price': prices.min(), 'max_price': prices.max(), 'mean_price': prices.mean(), 'latest_price': prices.iloc[-1], 'price_range_pct': (prices.max() - prices.min()) / prices.mean() * 100 if prices.mean() else 0.0}, 'data_quality': {'total_observations': len(prices), 'missing_values': prices.isna().sum(), 'time_span_hours': float((prices.index[-1] - prices.index[0]).total_seconds() / 3600.0) if len(prices) > 1 else 0.0}}
    return results

def _collect_levels_for_symbol(symbol_df: pd.DataFrame, side: str, max_levels: int = DEFAULT_MAX_LEVELS) -> Dict[int, Dict[str, list]]:
    out = {}
    for _, row in symbol_df.iterrows():
        base_price = _normalize_number(row['base_price'])
        if side not in row or pd.isna(row[side]):
            continue
        try:
            levels = parse_orderbook_blob(row[side])
        except Exception:
            continue
        for i, lvl in enumerate(levels[:max_levels], start=1):
            try:
                px, sz = (float(lvl['px']), float(lvl['sz']))
            except Exception:
                continue
            spread_distance = abs((px / base_price - 1.0) * 100.0)
            usd = px * sz
            bucket = out.setdefault(i, {'spread_distance': [], 'qty': [], 'usd': []})
            bucket['spread_distance'].append(spread_distance)
            bucket['qty'].append(sz)
            bucket['usd'].append(usd)
    return out

def _percentile(series: List[float], q: float) -> float:
    return float(pd.Series(series).quantile(q)) if series else float('nan')

def _mean(series: List[float]) -> float:
    return float(pd.Series(series).mean()) if series else float('nan')

def build_percentile_depth(results: Dict, percentiles: List[int], max_levels: int = DEFAULT_MAX_LEVELS) -> Dict[int, pd.DataFrame]:
    percentile_dfs = {}
    for perc in percentiles:
        q_spread = (100 - perc) / 100.0
        q_volume = perc / 100.0
        rows = []
        for symbol, blob in results.items():
            sdf = blob['df']
            for side in ('bids', 'asks'):
                buckets = _collect_levels_for_symbol(sdf, side, max_levels)
                acc_qty = acc_usd = 0.0
                for lvl in range(1, max_levels + 1):
                    if lvl in buckets:
                        b = buckets[lvl]
                        spread_p = _percentile(b['spread_distance'], q_spread)
                        qty_p = _percentile(b['qty'], q_volume) 
                        usd_p = _percentile(b['usd'], q_volume)
                        obs_count = len(b['qty'])
                    else:
                        spread_p = qty_p = usd_p = np.nan
                        obs_count = 0
                    
                    if not np.isnan(qty_p):
                        acc_qty += qty_p
                    if not np.isnan(usd_p):
                        acc_usd += usd_p
                    
                    rows.append({'symbol': symbol, 'side': 'bid' if side == 'bids' else 'ask', 'level': lvl, 'spread_distance': round(spread_p, 3) if not np.isnan(spread_p) else np.nan, 'qty': round(qty_p, 6) if not np.isnan(qty_p) else np.nan, 'usd': round(usd_p, 2) if not np.isnan(usd_p) else np.nan, 'acc_qty': round(acc_qty, 6) if not np.isnan(acc_qty) else np.nan, 'acc_usd': round(acc_usd, 2) if not np.isnan(acc_usd) else np.nan, 'obs_count': obs_count})
        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values(['symbol', 'side', 'level']).reset_index(drop=True)
        percentile_dfs[perc] = df
    return percentile_dfs

def build_mean_depth(results: Dict, max_levels: int = DEFAULT_MAX_LEVELS) -> pd.DataFrame:
    rows = []
    for symbol, blob in results.items():
        sdf = blob['df']
        for side in ('bids', 'asks'):
            buckets = _collect_levels_for_symbol(sdf, side, max_levels)
            acc_qty = acc_usd = 0.0
            for lvl in range(1, max_levels + 1):
                if lvl in buckets:
                    b = buckets[lvl]
                    spread_m, qty_m, usd_m = (_mean(b['spread_distance']), _mean(b['qty']), _mean(b['usd']))
                    obs_count = len(b['qty'])
                else:
                    spread_m = qty_m = usd_m = np.nan
                    obs_count = 0
                
                if not np.isnan(qty_m):
                    acc_qty += qty_m
                if not np.isnan(usd_m):
                    acc_usd += usd_m
                    
                rows.append({'symbol': symbol, 'side': 'bid' if side == 'bids' else 'ask', 'level': lvl, 'spread_distance': round(spread_m, 3) if not np.isnan(spread_m) else np.nan, 'qty': round(qty_m, 6) if not np.isnan(qty_m) else np.nan, 'usd': round(usd_m, 2) if not np.isnan(usd_m) else np.nan, 'acc_qty': round(acc_qty, 6) if not np.isnan(acc_qty) else np.nan, 'acc_usd': round(acc_usd, 2) if not np.isnan(acc_usd) else np.nan, 'obs_count': obs_count})
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['symbol', 'side', 'level']).reset_index(drop=True)
    return df

def calculate_stability_metrics(results: Dict, max_levels: int = DEFAULT_MAX_LEVELS) -> Dict:
    stability_data = {}
    
    for symbol, blob in results.items():
        sdf = blob['df'].copy()
        if 'ts' in sdf.columns:
            sdf = sdf.sort_values('ts')
        
        stability_data[symbol] = {}
        
        for side in ('bids', 'asks'):
            side_name = 'bid' if side == 'bids' else 'ask'
            stability_data[symbol][side_name] = {}
            
            level_qty_series = {}
            level_usd_series = {}
            
            for level in range(1, max_levels + 1):
                level_qty_series[level] = []
                level_usd_series[level] = []
            
            for _, row in sdf.iterrows():
                base_price = _normalize_number(row['base_price'])
                
                for level in range(1, max_levels + 1):
                    level_qty_series[level].append(np.nan)
                    level_usd_series[level].append(np.nan)
                
                if side in row and not pd.isna(row[side]):
                    try:
                        levels = parse_orderbook_blob(row[side])
                        acc_qty = 0.0
                        acc_usd = 0.0
                        
                        for i, lvl in enumerate(levels[:max_levels], start=1):
                            try:
                                px, sz = (float(lvl['px']), float(lvl['sz']))
                                acc_qty += sz
                                acc_usd += px * sz
                                level_qty_series[i][-1] = acc_qty
                                level_usd_series[i][-1] = acc_usd
                            except Exception:
                                continue
                    except Exception:
                        pass
            
            for level in range(1, max_levels + 1):
                qty_series = pd.Series(level_qty_series[level]).dropna()
                usd_series = pd.Series(level_usd_series[level]).dropna()
                
                qty_stats = {}
                usd_stats = {}
                
                if len(qty_series) > 1:
                    qty_stats = {
                        'sd': qty_series.std(),
                        'mean': qty_series.mean(),
                        'cv': qty_series.std() / qty_series.mean() if qty_series.mean() != 0 else np.nan,
                        'iqr': qty_series.quantile(0.75) - qty_series.quantile(0.25),
                        'autocorr': qty_series.autocorr(lag=1) if len(qty_series) > 2 else np.nan,
                        'count': len(qty_series)
                    }
                else:
                    qty_stats = {'sd': np.nan, 'mean': np.nan, 'cv': np.nan, 'iqr': np.nan, 'autocorr': np.nan, 'count': len(qty_series)}
                
                if len(usd_series) > 1:
                    usd_stats = {
                        'sd': usd_series.std(),
                        'mean': usd_series.mean(),
                        'cv': usd_series.std() / usd_series.mean() if usd_series.mean() != 0 else np.nan,
                        'iqr': usd_series.quantile(0.75) - usd_series.quantile(0.25),
                        'autocorr': usd_series.autocorr(lag=1) if len(usd_series) > 2 else np.nan,
                        'count': len(usd_series)
                    }
                else:
                    usd_stats = {'sd': np.nan, 'mean': np.nan, 'cv': np.nan, 'iqr': np.nan, 'autocorr': np.nan, 'count': len(usd_series)}
                
                stability_data[symbol][side_name][level] = {
                    'qty': qty_stats,
                    'usd': usd_stats
                }
    
    return stability_data

def calculate_volume_by_spread_comparison(results: Dict, percentiles: List[int], percentile_dfs: Dict, mean_df: pd.DataFrame, max_levels: int = DEFAULT_MAX_LEVELS) -> Dict:
    volume_data = {}
    symbols = list(results.keys())
    vol_keys = ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']
    
    for perc in percentiles:
        volume_data[f'p{perc}'] = {}
        spread_df = percentile_dfs[perc]
        usd_df = percentile_dfs[perc]
        
        for symbol in symbols:
            volume_data[f'p{perc}'][symbol] = {}
            
            for vol_key in vol_keys:
                volatility_pct = results[symbol]['volatility_metrics'][vol_key] * 100.0
                total_volume = 0.0
                
                symbol_spread = spread_df[spread_df['symbol'] == symbol].copy()
                symbol_usd = usd_df[usd_df['symbol'] == symbol].copy()
                
                if symbol_spread.empty or symbol_usd.empty:
                    volume_data[f'p{perc}'][symbol][vol_key] = 0.0
                    continue
                
                prev_level_spread = 0.0
                prev_level_volume = 0.0
                first_level_processed = False
                
                for level in range(1, max_levels + 1):
                    ask_spread = symbol_spread[(symbol_spread['side'] == 'ask') & (symbol_spread['level'] == level)]
                    bid_spread = symbol_spread[(symbol_spread['side'] == 'bid') & (symbol_spread['level'] == level)]
                    
                    ask_spread_val = ask_spread['spread_distance'].iloc[0] if len(ask_spread) > 0 and not pd.isna(ask_spread['spread_distance'].iloc[0]) else np.nan
                    bid_spread_val = bid_spread['spread_distance'].iloc[0] if len(bid_spread) > 0 and not pd.isna(bid_spread['spread_distance'].iloc[0]) else np.nan
                    
                    if not np.isnan(ask_spread_val) and not np.isnan(bid_spread_val):
                        level_spread = (ask_spread_val + bid_spread_val) / 2
                    elif not np.isnan(ask_spread_val):
                        level_spread = ask_spread_val
                    elif not np.isnan(bid_spread_val):
                        level_spread = bid_spread_val
                    else:
                        continue
                    
                    ask_usd = symbol_usd[(symbol_usd['side'] == 'ask') & (symbol_usd['level'] == level)]
                    bid_usd = symbol_usd[(symbol_usd['side'] == 'bid') & (symbol_usd['level'] == level)]
                    
                    ask_usd_val = ask_usd['acc_usd'].iloc[0] if len(ask_usd) > 0 and not pd.isna(ask_usd['acc_usd'].iloc[0]) else np.nan
                    bid_usd_val = bid_usd['acc_usd'].iloc[0] if len(bid_usd) > 0 and not pd.isna(bid_usd['acc_usd'].iloc[0]) else np.nan
                    
                    if not np.isnan(ask_usd_val) and not np.isnan(bid_usd_val):
                        level_volume = (ask_usd_val + bid_usd_val) / 2
                    elif not np.isnan(ask_usd_val):
                        level_volume = ask_usd_val
                    elif not np.isnan(bid_usd_val):
                        level_volume = bid_usd_val
                    else:
                        continue
                    
                    if level == 1 and not first_level_processed:
                        if volatility_pct < level_spread:
                            interpolation_factor = volatility_pct / level_spread if level_spread > 0 else 0
                            total_volume += level_volume * interpolation_factor
                            first_level_processed = True
                            break
                        else:
                            total_volume += level_volume
                            first_level_processed = True
                    else:
                        if volatility_pct >= level_spread:
                            total_volume += level_volume
                        else:
                            if level > 1 and prev_level_spread < volatility_pct < level_spread:
                                spread_range = level_spread - prev_level_spread
                                vol_offset = volatility_pct - prev_level_spread
                                interpolation_factor = vol_offset / spread_range if spread_range > 0 else 0
                                interpolated_volume = prev_level_volume + (level_volume - prev_level_volume) * interpolation_factor
                                total_volume += interpolated_volume
                            break
                    
                    prev_level_spread = level_spread
                    prev_level_volume = level_volume
                
                volume_data[f'p{perc}'][symbol][vol_key] = total_volume
    
    volume_data['mean'] = {}
    for symbol in symbols:
        volume_data['mean'][symbol] = {}
        
        for vol_key in vol_keys:
            volatility_pct = results[symbol]['volatility_metrics'][vol_key] * 100.0
            total_volume = 0.0
            
            symbol_spread = mean_df[mean_df['symbol'] == symbol].copy()
            symbol_usd = mean_df[mean_df['symbol'] == symbol].copy()
            
            if symbol_spread.empty or symbol_usd.empty:
                volume_data['mean'][symbol][vol_key] = 0.0
                continue
            
            prev_level_spread = 0.0
            prev_level_volume = 0.0
            first_level_processed = False
            
            for level in range(1, max_levels + 1):
                ask_spread = symbol_spread[(symbol_spread['side'] == 'ask') & (symbol_spread['level'] == level)]
                bid_spread = symbol_spread[(symbol_spread['side'] == 'bid') & (symbol_spread['level'] == level)]
                
                ask_spread_val = ask_spread['spread_distance'].iloc[0] if len(ask_spread) > 0 and not pd.isna(ask_spread['spread_distance'].iloc[0]) else np.nan
                bid_spread_val = bid_spread['spread_distance'].iloc[0] if len(bid_spread) > 0 and not pd.isna(bid_spread['spread_distance'].iloc[0]) else np.nan
                
                if not np.isnan(ask_spread_val) and not np.isnan(bid_spread_val):
                    level_spread = (ask_spread_val + bid_spread_val) / 2
                elif not np.isnan(ask_spread_val):
                    level_spread = ask_spread_val
                elif not np.isnan(bid_spread_val):
                    level_spread = bid_spread_val
                else:
                    continue
                
                ask_usd = symbol_usd[(symbol_usd['side'] == 'ask') & (symbol_usd['level'] == level)]
                bid_usd = symbol_usd[(symbol_usd['side'] == 'bid') & (symbol_usd['level'] == level)]
                
                ask_usd_val = ask_usd['acc_usd'].iloc[0] if len(ask_usd) > 0 and not pd.isna(ask_usd['acc_usd'].iloc[0]) else np.nan
                bid_usd_val = bid_usd['acc_usd'].iloc[0] if len(bid_usd) > 0 and not pd.isna(bid_usd['acc_usd'].iloc[0]) else np.nan
                
                if not np.isnan(ask_usd_val) and not np.isnan(bid_usd_val):
                    level_volume = (ask_usd_val + bid_usd_val) / 2
                elif not np.isnan(ask_usd_val):
                    level_volume = ask_usd_val
                elif not np.isnan(bid_usd_val):
                    level_volume = bid_usd_val
                else:
                    continue
                
                if level == 1 and not first_level_processed:
                    if volatility_pct < level_spread:
                        interpolation_factor = volatility_pct / level_spread if level_spread > 0 else 0
                        total_volume += level_volume * interpolation_factor
                        first_level_processed = True
                        break
                    else:
                        total_volume += level_volume
                        first_level_processed = True
                else:
                    if volatility_pct >= level_spread:
                        total_volume += level_volume
                    else:
                        if level > 1 and prev_level_spread < volatility_pct < level_spread:
                            spread_range = level_spread - prev_level_spread
                            vol_offset = volatility_pct - prev_level_spread
                            interpolation_factor = vol_offset / spread_range if spread_range > 0 else 0
                            interpolated_volume = prev_level_volume + (level_volume - prev_level_volume) * interpolation_factor
                            total_volume += interpolated_volume
                        break
                
                prev_level_spread = level_spread
                prev_level_volume = level_volume
            
            volume_data['mean'][symbol][vol_key] = total_volume
    
    return volume_data

def get_date_range_info(results: Dict) -> Dict:
    all_timestamps = []
    for symbol, data in results.items():
        df = data['df']
        if 'ts' in df.columns:
            timestamps = df['ts'].dropna()
            all_timestamps.extend(timestamps.tolist())
    
    if not all_timestamps:
        return {'initial_date': None, 'last_date': None, 'time_span_hours': 0.0, 'num_days': 0.0}
    
    all_timestamps = pd.to_datetime(all_timestamps)
    initial_date = all_timestamps.min()
    last_date = all_timestamps.max()
    time_span_hours = (last_date - initial_date).total_seconds() / 3600.0
    num_days = time_span_hours / 24.0
    
    return {
        'initial_date': initial_date,
        'last_date': last_date, 
        'time_span_hours': time_span_hours,
        'num_days': num_days
    }

def filter_empty_symbols(results: Dict, volume_data: Dict) -> Dict:
    """Remove symbols that have no valid data in volume tables"""
    symbols_to_remove = set()
    
    for table_key in volume_data.keys():
        for symbol in list(volume_data[table_key].keys()):
            symbol_data = volume_data[table_key][symbol]
            has_data = any(v > 0 for v in symbol_data.values() if not pd.isna(v))
            if not has_data:
                symbols_to_remove.add(symbol)
    
    for symbol in symbols_to_remove:
        if symbol in results:
            del results[symbol]
        for table_key in volume_data.keys():
            if symbol in volume_data[table_key]:
                del volume_data[table_key][symbol]
    
    return results

def add_summary_statistics(ws, data_dict: Dict, start_row: int, col_start: int = 2) -> int:
    """Add quantile summary rows for tables"""
    symbols = list(data_dict.keys())
    if not symbols:
        return start_row
    
    vol_keys = ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']
    quantile_labels = ['Min', '25%', '50%', '75%', 'Max', 'Mean']
    
    for i, label in enumerate(quantile_labels):
        row = start_row + i
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, color='0066CC')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, vol_key in enumerate(vol_keys):
            col = col_start + col_idx
            
            if vol_key in ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']:
                values = [data_dict[sym]['volatility_metrics'][vol_key] for sym in symbols if vol_key in data_dict[sym]['volatility_metrics']]
            else:
                values = [data_dict[sym][vol_key] for sym in symbols if vol_key in data_dict[sym]]
            
            if values:
                if label == 'Min':
                    stat_val = min(values)
                elif label == '25%':
                    stat_val = np.percentile(values, 25)
                elif label == '50%':
                    stat_val = np.percentile(values, 50)
                elif label == '75%':
                    stat_val = np.percentile(values, 75)
                elif label == 'Max':
                    stat_val = max(values)
                elif label == 'Mean':
                    stat_val = np.mean(values)
                
                cell = ws.cell(row=row, column=col, value=stat_val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if vol_key in ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']:
                    cell.style = 'percentage'
    
    return start_row + len(quantile_labels)

def create_orderbook_depth_plots(results: Dict, percentiles: List[int], max_levels: int = DEFAULT_MAX_LEVELS) -> List[str]:
    percentile_dfs = build_percentile_depth(results, percentiles, max_levels)
    mean_df = build_mean_depth(results, max_levels)
    symbols = list(results.keys())
    colors = plt.cm.Set1(np.linspace(0, 1, len(symbols)))
    plot_files = []
    
    for perc in sorted(percentiles):
        filename = PLOTS_DIR / f'orderbook_depth_p{perc:02d}.png'
        plot_files.append(filename)
        plt.figure(figsize=(12, 8))
        pct_df = percentile_dfs[perc]
        
        for i, symbol in enumerate(symbols):
            color = colors[i]
            ask_data = pct_df[(pct_df['symbol'] == symbol) & (pct_df['side'] == 'ask')].sort_values('level')
            bid_data = pct_df[(pct_df['symbol'] == symbol) & (pct_df['side'] == 'bid')].sort_values('level')
            
            ask_levels = []
            ask_values = []
            for level in range(1, max_levels + 1):
                level_data = ask_data[ask_data['level'] == level]
                if len(level_data) > 0:
                    ask_levels.append(level)
                    value = level_data['acc_usd'].iloc[0]
                    ask_values.append(max(value, 1) if not pd.isna(value) else 1)
            
            bid_levels = []
            bid_values = []
            for level in range(1, max_levels + 1):
                level_data = bid_data[bid_data['level'] == level]
                if len(level_data) > 0:
                    bid_levels.append(level)
                    value = level_data['acc_usd'].iloc[0]
                    bid_values.append(max(value, 1) if not pd.isna(value) else 1)
            
            if ask_levels and ask_values:
                plt.plot(ask_levels, ask_values, color=color, linewidth=2, linestyle='-', label=f'{symbol} Ask', marker='o', markersize=6)
            if bid_levels and bid_values:
                plt.plot(bid_levels, bid_values, color=color, linewidth=2, linestyle='--', label=f'{symbol} Bid', marker='s', markersize=6)
        
        plt.yscale('log')
        plt.xlabel('Orderbook Level', fontsize=12, fontweight='bold')
        plt.ylabel('Accumulated USD Value (Log Scale)', fontsize=12, fontweight='bold')
        plt.title(f'Orderbook Depth - Percentile {perc}% (Accumulated USD)', fontsize=14, fontweight='bold')
        plt.grid(True, alpha=0.3, which='both')
        plt.xticks(range(1, max_levels + 1), [f'N{i}' for i in range(1, max_levels + 1)])
        ax = plt.gca()
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}' if x >= 1000 else f'${x:.0f}'))
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
    
    filename = PLOTS_DIR / 'orderbook_depth_mean.png'
    plot_files.append(filename)
    plt.figure(figsize=(12, 8))
    
    for i, symbol in enumerate(symbols):
        color = colors[i]
        ask_data = mean_df[(mean_df['symbol'] == symbol) & (mean_df['side'] == 'ask')].sort_values('level')
        bid_data = mean_df[(mean_df['symbol'] == symbol) & (mean_df['side'] == 'bid')].sort_values('level')
        
        ask_levels = []
        ask_values = []
        for level in range(1, max_levels + 1):
            level_data = ask_data[ask_data['level'] == level]
            if len(level_data) > 0:
                ask_levels.append(level)
                value = level_data['acc_usd'].iloc[0]
                ask_values.append(max(value, 1) if not pd.isna(value) else 1)
        
        bid_levels = []
        bid_values = []
        for level in range(1, max_levels + 1):
            level_data = bid_data[bid_data['level'] == level]
            if len(level_data) > 0:
                bid_levels.append(level)
                value = level_data['acc_usd'].iloc[0]
                bid_values.append(max(value, 1) if not pd.isna(value) else 1)
        
        if ask_levels and ask_values:
            plt.plot(ask_levels, ask_values, color=color, linewidth=2, linestyle='-', label=f'{symbol} Ask', marker='o', markersize=6)
        if bid_levels and bid_values:
            plt.plot(bid_levels, bid_values, color=color, linewidth=2, linestyle='--', label=f'{symbol} Bid', marker='s', markersize=6)
    
    plt.yscale('log')
    plt.xlabel('Orderbook Level', fontsize=12, fontweight='bold')
    plt.ylabel('Accumulated USD Value (Log Scale)', fontsize=12, fontweight='bold')
    plt.title('Orderbook Depth - Mean Values (Accumulated USD)', fontsize=14, fontweight='bold')
    plt.grid(True, alpha=0.3, which='both')
    plt.xticks(range(1, max_levels + 1), [f'N{i}' for i in range(1, max_levels + 1)])
    ax = plt.gca()
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}' if x >= 1000 else f'${x:.0f}'))
    plt.tight_layout()
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.close()
    
    return plot_files

def create_excel_tables(results: Dict, percentiles: List[int], max_levels: int = DEFAULT_MAX_LEVELS, output_filename: str = EXCEL_FILENAME) -> None:
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = 'Trading Analysis'
    
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_font = Font(bold=True, size=14, color='000000')
    center = Alignment(horizontal='center', vertical='center')
    
    currency_style = NamedStyle(name="currency", number_format='[$-409]#,##0.00')
    percentage_style = NamedStyle(name="percentage", number_format='0.000%')
    
    if "currency" not in workbook.named_styles:
        workbook.add_named_style(currency_style)
    if "percentage" not in workbook.named_styles:
        workbook.add_named_style(percentage_style)

    symbols = list(results.keys())
    time_labels = ['3min', '5min', '10min', '30min', '60min', '90min', '1day', '1year']
    vol_keys = ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']
    
    date_info = get_date_range_info(results)
    
    row = 1
    
    if date_info['initial_date'] and date_info['last_date']:
        ws.cell(row=row, column=1, value=f"Initial Date: {date_info['initial_date'].strftime('%d/%m/%Y %H:%M')}").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value=f"Last Date: {date_info['last_date'].strftime('%d/%m/%Y %H:%M')}").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value=f"Time Span: {date_info['num_days']:.1f} days ({date_info['time_span_hours']:.1f} hours)").font = Font(bold=True, size=12)
        row += 2
    
    row += 1
    
    # Table 1: Volatility Analysis
    ws.cell(row=row, column=1, value='1. Volatility Analysis - Delta Percentages + Leverage Info').font = title_font
    table1_start = row + 1
    row += 2
    
    # Add summary statistics for Table 1
    row = add_summary_statistics(ws, results, row, 2)
    row += 1
    
    headers = ['Symbol'] + ['1 day' if tl == '1day' else '1 year' if tl == '1year' else tl for tl in time_labels] + ['Real Max Leverage', 'Max Quantity', 'Second Leverage']
    for c, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1
    
    for sym in symbols:
        ws.cell(row=row, column=1, value=sym).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = center
        for c, vk in enumerate(vol_keys, start=2):
            v = results[sym]['volatility_metrics'][vk]
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = center
            cell.style = percentage_style
        lev_info = results[sym]['leverage_info']
        tiers = lev_info.get('tiers', [])
        real_max_lev = lev_info.get('max_leverage', 0)
        ws.cell(row=row, column=len(headers) - 2, value=real_max_lev).alignment = center
        if len(tiers) >= 2:
            max_quantity = tiers[0][1]
            second_leverage = tiers[1][2]
            if max_quantity == float('inf'):
                max_qty_val = None
            else:
                max_qty_val = max_quantity
            cell = ws.cell(row=row, column=len(headers) - 1, value=max_qty_val)
            cell.alignment = center
            if max_qty_val:
                cell.style = currency_style
            ws.cell(row=row, column=len(headers), value=second_leverage).alignment = center
        else:
            ws.cell(row=row, column=len(headers) - 1, value=None).alignment = center
            ws.cell(row=row, column=len(headers), value=None).alignment = center
        row += 1
    
    table1_end = row - 1
    row += 2
    
    # Table 2: Spread Distances
    ws.cell(row=row, column=1, value="2. Spread Distances by Percentile + Mean").font = title_font
    table2_start = row + 1
    row += 2

    percentile_dfs = build_percentile_depth(results, percentiles, max_levels)
    mean_df = build_mean_depth(results, max_levels)
    
    headers = ["Token", "Level"]
    for perc in percentiles:
        headers.append(f"p{perc}")
    headers.append("mean")

    for c, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1

    for symbol in symbols:
        for level in range(1, max_levels + 1):
            col = 1
            ws.cell(row=row, column=col, value=symbol).alignment = center
            col += 1
            ws.cell(row=row, column=col, value=f"N{level}").alignment = center
            col += 1

            for perc in percentiles:
                pct_df = percentile_dfs[perc]
                ask_val = pct_df.loc[(pct_df['symbol'] == symbol) & (pct_df['side'] == 'ask') & (pct_df['level'] == level), 'spread_distance']
                bid_val = pct_df.loc[(pct_df['symbol'] == symbol) & (pct_df['side'] == 'bid') & (pct_df['level'] == level), 'spread_distance']
                
                ask_num = ask_val.iloc[0] if len(ask_val) > 0 and not pd.isna(ask_val.iloc[0]) else np.nan
                bid_num = bid_val.iloc[0] if len(bid_val) > 0 and not pd.isna(bid_val.iloc[0]) else np.nan
                
                if not np.isnan(ask_num) and not np.isnan(bid_num):
                    mean_val = (ask_num + bid_num) / 2 / 100.0
                elif not np.isnan(ask_num):
                    mean_val = ask_num / 100.0
                elif not np.isnan(bid_num):
                    mean_val = bid_num / 100.0
                else:
                    mean_val = None
                
                cell = ws.cell(row=row, column=col, value=mean_val)
                cell.alignment = center
                if mean_val is not None:
                    cell.style = percentage_style
                col += 1

            ask_val = mean_df.loc[(mean_df['symbol'] == symbol) & (mean_df['side'] == 'ask') & (mean_df['level'] == level), 'spread_distance']
            bid_val = mean_df.loc[(mean_df['symbol'] == symbol) & (mean_df['side'] == 'bid') & (mean_df['level'] == level), 'spread_distance']
            
            ask_num = ask_val.iloc[0] if len(ask_val) > 0 and not pd.isna(ask_val.iloc[0]) else np.nan
            bid_num = bid_val.iloc[0] if len(bid_val) > 0 and not pd.isna(bid_val.iloc[0]) else np.nan
            
            if not np.isnan(ask_num) and not np.isnan(bid_num):
                mean_val = (ask_num + bid_num) / 2 / 100.0
            elif not np.isnan(ask_num):
                mean_val = ask_num / 100.0
            elif not np.isnan(bid_num):
                mean_val = bid_num / 100.0
            else:
                mean_val = None
                
            cell = ws.cell(row=row, column=col, value=mean_val)
            cell.alignment = center
            if mean_val is not None:
                cell.style = percentage_style

            row += 1
    
    table2_end = row - 1
    row += 2

    # Table 3: Token Quantities with Stability Analysis
    stability_data = calculate_stability_metrics(results, max_levels)
    
    ws.cell(row=row, column=1, value='3. Token Quantities Stability Analysis').font = title_font
    table3_start = row + 1
    row += 2
    
    stability_headers = ["Symbol", "Level", "Side", "Mean", "SD", "CV", "IQR", "Autocorr", "Count"]
    for c, hdr in enumerate(stability_headers, start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1
    
    for symbol in symbols:
        for level in range(1, max_levels + 1):
            for side in ['bid', 'ask']:
                if symbol in stability_data and side in stability_data[symbol] and level in stability_data[symbol][side]:
                    stats = stability_data[symbol][side][level]['qty']
                    col = 1
                    ws.cell(row=row, column=col, value=symbol).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=f"N{level}").alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=side.upper()).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['mean']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['sd']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['cv']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['iqr']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['autocorr']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['count']).alignment = center
                    row += 1
    
    table3_end = row - 1
    row += 2

    # Table 4: USD Values Stability Analysis
    ws.cell(row=row, column=1, value='4. USD Values Stability Analysis').font = title_font
    table4_start = row + 1
    row += 2
    
    for c, hdr in enumerate(stability_headers, start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    row += 1
    
    for symbol in symbols:
        for level in range(1, max_levels + 1):
            for side in ['bid', 'ask']:
                if symbol in stability_data and side in stability_data[symbol] and level in stability_data[symbol][side]:
                    stats = stability_data[symbol][side][level]['usd']
                    col = 1
                    ws.cell(row=row, column=col, value=symbol).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=f"N{level}").alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=side.upper()).alignment = center
                    col += 1
                    cell = ws.cell(row=row, column=col, value=stats['mean'])
                    cell.alignment = center
                    if stats['mean'] and not pd.isna(stats['mean']):
                        cell.style = currency_style
                    col += 1
                    cell = ws.cell(row=row, column=col, value=stats['sd'])
                    cell.alignment = center
                    if stats['sd'] and not pd.isna(stats['sd']):
                        cell.style = currency_style
                    col += 1
                    ws.cell(row=row, column=col, value=stats['cv']).alignment = center
                    col += 1
                    cell = ws.cell(row=row, column=col, value=stats['iqr'])
                    cell.alignment = center
                    if stats['iqr'] and not pd.isna(stats['iqr']):
                        cell.style = currency_style
                    col += 1
                    ws.cell(row=row, column=col, value=stats['autocorr']).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=stats['count']).alignment = center
                    row += 1
    
    table4_end = row - 1
    row += 2

    # Calculate volume data
    volume_data = calculate_volume_by_spread_comparison(results, percentiles, percentile_dfs, mean_df, max_levels)
    
    # Filter empty symbols
    results = filter_empty_symbols(results, volume_data)
    symbols = list(results.keys())
    
    # Volume tables with summary statistics
    volume_tables = [f'p{p}' for p in percentiles] + ['mean']
    table_starts = []
    table_ends = []
    
    for i, table_key in enumerate(volume_tables):
        table_num = 5 + i
        ws.cell(row=row, column=1, value=f'{table_num}. Max Investment Volume - {table_key.upper()}').font = title_font
        table_start = row + 1
        table_starts.append(table_start)
        row += 2
        
        # Add summary statistics
        volume_summary_data = {}
        for sym in symbols:
            volume_summary_data[sym] = volume_data[table_key][sym]
        
        row = add_summary_statistics_volume(ws, volume_summary_data, row, 2)
        row += 1
        
        headers = ['Symbol'] + ['1 day' if tl == '1day' else '1 year' if tl == '1year' else tl for tl in time_labels]
        for c, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        row += 1
        
        for sym in symbols:
            ws.cell(row=row, column=1, value=sym).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = center
            
            for c, vk in enumerate(vol_keys, start=2):
                volume = volume_data[table_key][sym][vk]
                cell = ws.cell(row=row, column=c, value=volume if volume > 0 else None)
                cell.alignment = center
                if volume > 0:
                    cell.style = currency_style
            row += 1
        
        table_end = row - 1
        table_ends.append(table_end)
        row += 2
    
    row += 1
    
    # Add plots section
    final_table_num = 5 + len(volume_tables)
    ws.cell(row=row, column=1, value=f'{final_table_num}. Orderbook Depth Analysis Plots').font = title_font
    row += 2
    
    depth_plot_files = create_orderbook_depth_plots(results, percentiles, max_levels)
    plot_row = row
    plots_per_row = 2
    plot_width_cells = 15
    plot_height_cells = 25
    
    for i, plot_file in enumerate(depth_plot_files):
        if Path(plot_file).exists():
            plot_col = 1 + i % plots_per_row * (plot_width_cells + 1)
            current_plot_row = plot_row + i // plots_per_row * (plot_height_cells + 2)
            plot_name = plot_file.name.replace('.png', '').replace('orderbook_depth_', '').upper()
            ws.cell(row=current_plot_row, column=plot_col, value=f'Plot: {plot_name}').font = Font(bold=True, size=12)
            img = Image(plot_file)
            img.width = 600
            img.height = 400
            img.anchor = f'{get_column_letter(plot_col)}{current_plot_row + 1}'
            ws.add_image(img)

    try:
        ws.row_dimensions.group(table1_start, table1_end, outline_level=1, hidden=False)
        ws.row_dimensions.group(table2_start, table2_end, outline_level=1, hidden=False)
        ws.row_dimensions.group(table3_start, table3_end, outline_level=1, hidden=False) 
        ws.row_dimensions.group(table4_start, table4_end, outline_level=1, hidden=False)
        for table_start, table_end in zip(table_starts, table_ends):
            ws.row_dimensions.group(table_start, table_end, outline_level=1, hidden=False)
    except:
        pass

    ws.column_dimensions['A'].width = 12
    for c in range(2, len(symbols) + 50):
        ws.column_dimensions[get_column_letter(c)].width = 12
    
    workbook.save(output_filename)

def add_summary_statistics_volume(ws, data_dict: Dict, start_row: int, col_start: int = 2) -> int:
    symbols = list(data_dict.keys())
    if not symbols:
        return start_row
    
    vol_keys = ['3min_vol', '5min_vol', '10min_vol', '30min_vol', '60min_vol', '90min_vol', '1day_vol', '1year_vol']
    quantile_labels = ['Min', '25%', '50%', '75%', 'Max', 'Mean']
    
    for i, label in enumerate(quantile_labels):
        row = start_row + i
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, color='0066CC')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, vol_key in enumerate(vol_keys):
            col = col_start + col_idx
            
            values = [data_dict[sym][vol_key] for sym in symbols if vol_key in data_dict[sym] and data_dict[sym][vol_key] > 0]
            
            if values:
                if label == 'Min':
                    stat_val = min(values)
                elif label == '25%':
                    stat_val = np.percentile(values, 25)
                elif label == '50%':
                    stat_val = np.percentile(values, 50)
                elif label == '75%':
                    stat_val = np.percentile(values, 75)
                elif label == 'Max':
                    stat_val = max(values)
                elif label == 'Mean':
                    stat_val = np.mean(values)
                
                cell = ws.cell(row=row, column=col, value=stat_val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.style = 'currency'
    
    return start_row + len(quantile_labels)

def parse_arguments():
    parser = argparse.ArgumentParser(description='Volatility and orderbook analysis with configurable levels and stability metrics')
    parser.add_argument('csv_file', help='Path to CSV file with orderbook data')
    parser.add_argument('symbols', nargs='*', help='Optional: specific symbols to analyze')
    parser.add_argument('--percentiles', nargs='+', type=int, default=DEFAULT_PERCENTILES, help=f'Percentiles to calculate (1-99). Default: {DEFAULT_PERCENTILES}')
    parser.add_argument('--sep', default=',', help='CSV separator/delimiter (default: ",")')
    parser.add_argument('--levels', type=int, default=DEFAULT_MAX_LEVELS, help=f'Maximum number of orderbook levels to process (default: {DEFAULT_MAX_LEVELS})')
    
    args = parser.parse_args()
    
    for p in args.percentiles:
        if not 1 <= p <= 99:
            parser.error(f'Percentile {p} must be between 1 and 99')
    args.percentiles = sorted(set(args.percentiles))
    
    if 50 not in args.percentiles:
        args.percentiles.append(50)
        args.percentiles = sorted(args.percentiles)
    
    if args.levels < 1:
        parser.error('Levels must be at least 1')
    
    return args

def main():
    args = parse_arguments()
    if not Path(args.csv_file).exists():
        raise SystemExit(f'CSV file not found: {args.csv_file}')
    if not Path(LEVERAGE_DATA_FILE).exists():
        print(f"Warning: Leverage data file not found: {LEVERAGE_DATA_FILE}")
    
    print(f"Processing with max {args.levels} orderbook levels")
    print(f"Total plots generated: {len(args.percentiles) + 1} (orderbook depth plots only)")
    
    results = analyze_csv(args.csv_file, args.symbols, sep=args.sep)
    create_excel_tables(results, args.percentiles, args.levels, output_filename=EXCEL_FILENAME)
    
    print(f"Analysis complete. Results saved to {EXCEL_FILENAME}")
    print(f"Plots saved to {PLOTS_DIR}")

if __name__ == '__main__':
    main()